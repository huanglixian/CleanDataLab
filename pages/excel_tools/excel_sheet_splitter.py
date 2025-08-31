import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import io
import zipfile
import shutil
import tempfile
import time
from datetime import datetime
from common.ui_style import apply_custom_style
from common.file_processing_queue import fp_queue


def split_excel_file(file_data, file_name, zip_writer=None):
    """核心拆分函数：使用文件复制方式完整保留所有样式"""
    base_name = Path(file_name).stem
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        temp_file.write(file_data.getvalue())
        temp_path = temp_file.name
    
    try:
        workbook = load_workbook(temp_path, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        # 如果只是获取sheet信息，直接返回
        if zip_writer is None:
            return len(sheet_names)
        
        if len(sheet_names) == 1:
            return 0  # 跳过单sheet文件
        
        # 为每个sheet创建独立文件
        for sheet_name in sheet_names:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as target_temp:
                shutil.copy2(temp_path, target_temp.name)
                
                wb = load_workbook(target_temp.name)
                for remove_name in [n for n in wb.sheetnames if n != sheet_name]:
                    wb.remove(wb[remove_name])
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                wb.close()
                
                zip_writer.writestr(f"{base_name}-{sheet_name}.xlsx", excel_buffer.getvalue())
        
        return len(sheet_names)  # 返回处理的sheet数量
        
    finally:
        Path(temp_path).unlink(missing_ok=True)




def process_files_batch(files_data):
    """批处理文件并返回结果"""
    zip_buffer = io.BytesIO()
    results = []
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for file_data, file_name in files_data:
            try:
                sheet_count = split_excel_file(file_data, file_name, zipf)
                if sheet_count == 0:
                    results.append([file_name, 1, '⏩ 已跳过(单sheet)'])
                else:
                    results.append([file_name, sheet_count, f'✅ 已拆分({sheet_count}个sheet)'])
            except Exception as e:
                results.append([file_name, 0, f'❌ {str(e)[:15]}...'])
    
    return zip_buffer, results


def main():
    st.set_page_config(page_title="Excel Sheet 拆分工具", page_icon="📄", layout="centered")
    apply_custom_style()
    
    # 初始化
    if "excel_split_key" not in st.session_state:
        st.session_state.excel_split_key = 0
    
    st.title("📄 Excel Sheet 拆分工具")
    st.markdown("将多个 Sheet 的 Excel 文件拆分为独立的 Excel 文件")
    st.markdown("---")
    
    uploaded_files = st.file_uploader(
        "请选择需要拆分的Excel文件（可多选）。**注意！只支持.xlsx格式！**",
        type=['xlsx'],
        accept_multiple_files=True,
        help="只支持.xlsx格式以确保样式完整保留，如有.xls文件请先转换为.xlsx",
        key=f"uploader_{st.session_state.excel_split_key}"
    )
    
    if uploaded_files:
        st.info(f"📊 已选择 {len(uploaded_files)} 个Excel文件")
        
        # Excel-Sheet预览
        st.subheader("📋 Excel-Sheet预览")
        file_info = []
        processable_count = 0
        
        for file in uploaded_files:
            sheet_count = split_excel_file(file, file.name)
            size = len(file.getvalue()) / 1024
            
            if sheet_count > 1:
                processable_count += 1
                status = f"🔄 将拆分({sheet_count}个sheet)"
            elif sheet_count == 1:
                status = "⏩ 将跳过(单sheet)"
            else:
                status = "❌ 读取失败"
                
            file_info.append([file.name, sheet_count, f"{size:.1f}", status])
        
        df = pd.DataFrame(file_info, columns=['文件名', 'Sheet数量', '大小(KB)', '状态'])
        st.dataframe(df, use_container_width=True, hide_index=True)
        st.success(f"其中 {processable_count} 个文件将被拆分")
        
        if st.button("🔄 开始拆分", type="primary", use_container_width=True, disabled="excel_split_task_running" in st.session_state):
            st.session_state.excel_split_task_running = True
            st.rerun()
        
        # 处理任务
        if st.session_state.get('excel_split_task_running') and not st.session_state.get('excel_split_result'):
            files_data = [(f, f.name) for f in uploaded_files]
            task_id = fp_queue.submit_task(files_data, process_files_batch)
            
            # 状态显示
            status_placeholder = st.empty()
            
            while True:
                position = fp_queue.get_task_position(task_id)
                task_status = fp_queue.get_task_status(task_id)
                
                if position > 0:
                    status_placeholder.warning(f"⏳ 排队中，第 {position} 位")
                elif task_status == "processing":
                    status_placeholder.info("🔄 正在拆分中...")
                elif task_status == "completed":
                    status_placeholder.success("✅ 拆分完成")
                    zip_buffer, results = fp_queue.wait_for_task(task_id)
                    break
                else:
                    zip_buffer, results = None, None
                    break
                
                time.sleep(1)
            
            if zip_buffer and results:
                st.session_state.excel_split_result = (zip_buffer, results)
                status_placeholder.empty()
            else:
                status_placeholder.error("拆分超时，请重试")
        
        # 显示结果
        if st.session_state.get('excel_split_result'):
            zip_buffer, results = st.session_state.excel_split_result
            st.success("✅ 拆分完成!")
            
            # 显示结果和统计
            st.dataframe(pd.DataFrame(results, columns=['文件名', 'Sheet数量', '状态']), use_container_width=True, hide_index=True)
            
            processed = sum(1 for r in results if r[2].startswith('✅'))
            skipped = sum(1 for r in results if r[2].startswith('⏩'))
            failed = sum(1 for r in results if r[2].startswith('❌'))
            
            col1, col2, col3 = st.columns(3)
            col1.metric("已拆分", processed)
            col2.metric("已跳过", skipped)
            col3.metric("失败", failed)
            
            # 下载
            if processed > 0:
                timestamp = datetime.now().strftime("%Y%m%d%H%M")
                filename = f"excel_sheet_拆分_{timestamp}.zip"
                
                col1, col2 = st.columns([2, 1])
                with col1:
                    st.download_button("📥 下载拆分文件", zip_buffer.getvalue(), filename, "application/zip", type="primary", use_container_width=True)
                with col2:
                    if st.button("🔄 重置页面", type="secondary", use_container_width=True):
                        st.session_state.excel_split_key += 1
                        st.session_state.pop('excel_split_result', None)
                        st.session_state.pop('excel_split_task_running', None)
                        st.rerun()


if __name__ == "__main__":
    main()