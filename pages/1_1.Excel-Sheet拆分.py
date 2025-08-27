import streamlit as st
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
import io
import zipfile
import glob
import shutil
import tempfile
from common.ui_style import apply_custom_style


def split_excel_file(file_data, file_name, output_folder_name, zip_writer):
    """核心拆分函数：使用文件复制方式完整保留所有样式"""
    base_name = Path(file_name).stem
    
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
        temp_file.write(file_data.read() if hasattr(file_data, 'read') else file_data.getvalue())
        temp_path = temp_file.name
    
    try:
        workbook = load_workbook(temp_path, data_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        
        if len(sheet_names) == 1:
            return 0  # 跳过
        
        # 为每个sheet创建独立文件
        for sheet_name in sheet_names:
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as target_temp:
                shutil.copy2(temp_path, target_temp.name)
                
                wb = load_workbook(target_temp.name)
                for remove_name in [n for n in wb.sheetnames if n != sheet_name]:
                    wb.remove(wb[remove_name])
                
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                wb.close()
                
                zip_writer.writestr(f"{output_folder_name}/{base_name}-{sheet_name}.xlsx", excel_buffer.getvalue())
        
        return len(sheet_names)  # 返回处理的sheet数量
        
    finally:
        Path(temp_path).unlink(missing_ok=True)


def get_sheet_count(file_path):
    """获取Excel文件的sheet数量"""
    try:
        wb = load_workbook(file_path, data_only=True)
        count = len(wb.sheetnames)
        wb.close()
        return count
    except Exception:
        return 0


def process_and_display(files_data, is_batch=False):
    """统一处理和显示逻辑"""
    if not files_data:
        return st.warning("📂 未找到Excel文件")
    
    # 分析文件信息
    file_info = []
    processable_count = 0
    
    for file_path, file_name in files_data:
        sheet_count = get_sheet_count(file_path if isinstance(file_path, str) else file_path)
        size = (Path(file_path).stat().st_size if isinstance(file_path, str) else len(file_path.getvalue())) / 1024
        
        if sheet_count > 1:
            processable_count += 1
            status = f"🔄 将拆分({sheet_count}个sheet)"
        elif sheet_count == 1:
            status = "⏩ 将跳过(单sheet)"
        else:
            status = "❌ 读取失败"
        
        file_info.append([file_name, sheet_count, f"{size:.1f}", status])
    
    # 显示文件信息
    st.subheader("📋 发现的Excel文件")
    df = pd.DataFrame(file_info, columns=['文件名', 'Sheet数量', '大小(KB)', '状态'])
    st.dataframe(df, use_container_width=True, hide_index=True)
    
    if is_batch:
        st.success(f"📊 共发现 {len(files_data)} 个文件，其中 {processable_count} 个将被拆分")
        if processable_count == 0:
            return st.warning("没有多个sheet的Excel文件需要拆分")
    else:
        st.success(f"发现 {file_info[0][1]} 个 Sheet")
    
    # 处理按钮和逻辑
    if st.button("🔄 开始批量拆分" if is_batch else "🔄 开始拆分", type="primary", use_container_width=True):
        # 确定输出文件夹名称
        if is_batch:
            first_file_path = files_data[0][0]
            source_folder = Path(first_file_path).parent.name if isinstance(first_file_path, str) else "批量"
            folder_name = f"{source_folder}_批量拆分"
        else:
            folder_name = f"{Path(files_data[0][1]).stem}_sheet拆分"
        
        # 处理文件
        with st.spinner("正在处理Excel文件..."):
            zip_buffer = io.BytesIO()
            results = []
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for file_path, file_name in files_data:
                    try:
                        if isinstance(file_path, str):
                            with open(file_path, 'rb') as f:
                                file_data = io.BytesIO(f.read())
                        else:
                            file_data = file_path
                        
                        sheet_count = split_excel_file(file_data, file_name, folder_name, zipf)
                        
                        if sheet_count == 0:
                            results.append([file_name, 1, '⏩ 已跳过(单sheet)'])
                        else:
                            results.append([file_name, sheet_count, f'✅ 已拆分({sheet_count}个sheet)'])
                            
                    except Exception as e:
                        results.append([file_name, 0, f'❌ {str(e)[:15]}...'])
        
        # 显示结果
        st.success("✅ 处理完成!")
        st.subheader("📦 处理结果")
        result_df = pd.DataFrame(results, columns=['源文件', 'Sheet数量', '状态'])
        st.dataframe(result_df, use_container_width=True, hide_index=True)
        
        # 统计信息
        processed = sum(1 for r in results if r[2].startswith('✅'))
        skipped = sum(1 for r in results if r[2].startswith('⏩'))
        failed = sum(1 for r in results if r[2].startswith('❌'))
        
        col1, col2, col3 = st.columns(3)
        col1.metric("已拆分", processed)
        col2.metric("已跳过", skipped)
        col3.metric("失败", failed)
        
        # 下载
        st.download_button(
            "📥 下载所有拆分文件",
            zip_buffer.getvalue(),
            f"{folder_name}.zip",
            "application/zip",
            type="primary",
            use_container_width=True
        )
        st.info(f"💡 所有拆分文件已统一放在 `{folder_name}` 文件夹中")


def main():
    st.set_page_config(page_title="Excel Sheet 拆分工具", page_icon="📄", layout="centered")
    apply_custom_style()
    
    st.title("📄 Excel Sheet 拆分工具")
    st.markdown("将多个 Sheet 的 Excel 文件拆分为独立的 Excel 文件")
    
    mode = st.radio(
        "选择操作模式",
        ["🔹 单文件拆分", "📁 批量文件夹拆分"],
        horizontal=True,
        help="批量模式会自动跳过单sheet文件，完整保留所有样式"
    )
    
    if mode == "📁 批量文件夹拆分":
        folder_path = st.text_input(
            "📁 请输入Excel文件夹路径",
            placeholder="例如: /Users/用户名/Documents/Excel文件夹",
            help="输入包含Excel文件的文件夹绝对路径，支持子文件夹递归搜索，自动跳过单sheet文件"
        )
        
        if folder_path:
            if Path(folder_path).exists():
                excel_files = []
                for ext in ['*.xlsx', '*.xls']:
                    excel_files.extend(glob.glob(f"{folder_path}/**/{ext}", recursive=True))
                
                files_data = [(file_path, Path(file_path).name) for file_path in excel_files]
                process_and_display(files_data, is_batch=True)
            else:
                st.error("❌ 文件夹路径不存在，请检查路径是否正确")

    else:  # 单文件模式
        uploaded_file = st.file_uploader(
            "选择要拆分的Excel文件", 
            type=['xlsx', 'xls'], 
            help="支持 .xlsx 和 .xls 格式，完整保留所有样式"
        )
        
        if uploaded_file:
            process_and_display([(uploaded_file, uploaded_file.name)], is_batch=False)


if __name__ == "__main__":
    main()